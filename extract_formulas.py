import openpyxl
import sys
import json

def extract():
    file_path = "excel1.xlsm"
    try:
        wb_formulas = openpyxl.load_workbook(file_path, data_only=False, keep_vba=True)
        if "PRINCIPLE 6" not in wb_formulas.sheetnames:
            print(f"Sheet 'PRINCIPLE 6' not found. Available sheets: {wb_formulas.sheetnames}")
            return
            
        sheet = wb_formulas["PRINCIPLE 6"]
        
        extracted_data = []
        for r_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=40), start=1):
            row_data = {"row_index": r_idx, "cells": []}
            has_content = False
            for col_idx, cell in enumerate(row, start=1):
                val = str(cell.value) if cell.value is not None else ""
                if val.strip() and "=" in val:
                    has_content = True
                    row_data["cells"].append({"col": cell.column_letter, "val": val})
                elif val.strip() and col_idx < 4:
                    has_content = True
                    row_data["cells"].append({"col": cell.column_letter, "val": str(val)[:50]})
            if has_content:
                extracted_data.append(row_data)
                
        with open("formulas_utf8.json", "w", encoding="utf-8") as f:
            json.dump(extracted_data, f, indent=2)
            
    except Exception as e:
        print(f"Python Error: {e}")

if __name__ == "__main__":
    extract()
